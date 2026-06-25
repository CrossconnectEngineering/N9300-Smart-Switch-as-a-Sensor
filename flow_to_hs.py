#!/usr/bin/env python3
"""
flow_to_hs.py — Convert NX-OS DPU flow dumps into Hypershield-style intent table.

USAGE:
        python flow_to_hs.py <capture_dir> -o intent.xlsx

INPUT:  a directory of flowcap_*.txt files written by flow_collector.py
OUTPUT: .xlsx with three columns for review/import:
        Source | Destination | PROTOCOL [PORTS]

REDUCTION PRINCIPLES (in order):
  1. STATEFUL.    Responder rows confirm flow completion (see #5), then are
                  dropped from the output. Only initiators emit policy rows.
  2. GLOBAL.      Drop all direction / interface / DPU context. Same flow seen
                  on DPU1 and DPU2 collapses to one tuple.
  3. UNORDERED.   No rule precedence. Output rows are unordered permits.
  4. PORT-TAGGED. Multiple ports between the same (src,dst) collapse into a
                  comma-list on a single row.
  5. CONFIRMED.   An initiator tuple (src, dst, proto, dport) is emitted only
                  if a matching responder row is observed in at least one
                  snapshot. Flows that stay init-only across the entire
                  capture window are evidence of blocked traffic and are
                  never emitted.

ROW-GROUPING RULE (deterministic, two-pass):
  Pass 1: Group by (destination, port-set) — sources sharing the exact same
          port-set against the same destination comma-collapse.
  Pass 2: Group by (source-set, port-set) — destinations sharing the exact
          same source-set and port-set comma-collapse.
  A given (source, destination) pair appears in exactly one output row.
"""

import argparse
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


PROTO_NAME = {6: "TCP", 17: "UDP", 1: "ICMP"}


def parse_dump(text):
    """
    Yield (src, dst, proto_num, sport, dport, role) tuples from a shflowdpu
    text dump.

    The dump format is space-separated with this column order:
      Session  Hash32b  Handle  Src-Ip  Dst-Ip  Proto  Src-Port  Dst-Port
      L2_L3  L2L3_Id  Role  Dir  PolId

    Section gating: only ingest flow rows that appear under a clean
    `shflowdpu1` or `shflowdpu2` command. Rows under filtered variants
    (e.g. `shflowdpu1 | grep ...`) are engineer debugging output and may
    represent stale state from earlier in the terminal session. They are
    skipped to prevent stale rows from contaminating the snapshot.
    """
    flow_re = re.compile(
        r"^\s*(\d+)\s+0x[0-9a-fA-F]+\s+\S+\s+"
        r"(\S+)\s+(\S+)\s+"               # src ip, dst ip
        r"(\d+)\s+(\d+)\s+(\d+)\s+"        # proto, sport, dport
        r"\S+\s+\S+\s+"                    # L2_L3, L2L3_Id
        r"(\S+)\s+"                        # Role
        r"(\S+)\s+(\S+)\s*$"               # Dir, PolId
    )
    # Clean dump command. Matches the team alias (`shflowdpu1`/`shflowdpu2`)
    # AND the underlying real NX-OS command (`slot 1 dpu N dpctl show flow`)
    # the flow_collector writes. Both must be unpiped — anything after them
    # (e.g. ` | grep ...`) means we're looking at filtered debug output and
    # should not be ingested as snapshot data.
    clean_dump_re = re.compile(
        r"(?:shflowdpu[1-4]|slot\s+\d+\s+dpu\s+\d+\s+dpctl\s+show\s+flow)\s*$"
    )
    # Any prompt line. When we see one that isn't a clean dump command, we
    # leave the ingest window.
    prompt_re = re.compile(r"^[A-Za-z0-9._-]+#\s")

    ingesting = False
    for raw in text.splitlines():
        line = raw.rstrip("\r")
        if prompt_re.match(line):
            # Toggle ingest based on whether this prompt issued a clean dump.
            ingesting = bool(clean_dump_re.search(line))
            continue
        if not ingesting:
            continue
        m = flow_re.match(line)
        if not m:
            continue
        _, src, dst, proto, sport, dport, role, _dir, _polid = m.groups()
        yield (src, dst, int(proto), int(sport), int(dport), role)


def reduce_flows(tuples):
    """
    Apply the reduction principles and return rows ready for emission.

    Returns: (rows, stats) where stats is a dict reporting filtering counts.

    rows: list of (sources_tuple, destinations_tuple, port_tags_tuple).
    """
    # Build two sets:
    #   - initiator_tuples: every (src, dst, proto, dport) that appears as initiator
    #   - confirmed_tuples: same key, but only where a matching responder row exists
    #
    # A responder row for initiator (A, B, proto, dport) flips both addresses and
    # ports, so it appears as (src=B, dst=A, proto=proto, sport=dport). We index
    # responders by (responder_src, responder_dst, proto, responder_sport) and
    # match that against the initiator's (dst, src, proto, dport).
    initiator_tuples = set()
    responder_tuples = set()
    for src, dst, proto, sport, dport, role in tuples:
        if role.lower() == "initiator":
            initiator_tuples.add((src, dst, proto, dport))
        elif role.lower() == "responder":
            # Index by the conversation key the responder represents,
            # i.e. the inverse of itself.
            responder_tuples.add((dst, src, proto, sport))

    confirmed_tuples = initiator_tuples & responder_tuples
    unconfirmed_tuples = initiator_tuples - responder_tuples

    # Principle 1: stateful — drop responders, initiators only
    # Principle 2: global — drop direction; collapse to unique 4-tuples
    # NEW: bidirectional confirmation — only keep initiators with matching responder
    # Build the (src,dst) -> {(proto, dport)} map from confirmed tuples only.
    src_dst_to_ports = defaultdict(set)
    for src, dst, proto, dport in confirmed_tuples:
        src_dst_to_ports[(src, dst)].add((proto, dport))

    # Principle 4: port-tagged — render each (proto, dport) as a tag string
    # and freeze the port-set per (src,dst) for grouping.
    def render_port_tags(port_pairs):
        # Sort: TCP before UDP, then by port number
        proto_order = {6: 0, 17: 1, 1: 2}
        sorted_pairs = sorted(
            port_pairs,
            key=lambda pp: (proto_order.get(pp[0], 99), pp[1]),
        )
        return tuple(
            f"{PROTO_NAME.get(p, str(p))} [{dp}]" for p, dp in sorted_pairs
        )

    src_dst_to_port_tags = {
        sd: render_port_tags(ports) for sd, ports in src_dst_to_ports.items()
    }

    # Group by (destination, frozen-port-set). All sources sharing the exact
    # same port-set against the same destination collapse to one row.
    group_key_to_sources = defaultdict(set)
    for (src, dst), port_tags in src_dst_to_port_tags.items():
        group_key = (dst, port_tags)
        group_key_to_sources[group_key].add(src)

    # Second-pass collapse: rows that share the exact same source-set and
    # port-set against different destinations collapse to one row with a
    # comma-list of destinations.
    pair_key_to_destinations = defaultdict(set)
    for (dst, port_tags), sources in group_key_to_sources.items():
        sources_frozen = tuple(sorted(sources, key=_ip_sort_key))
        pair_key = (sources_frozen, port_tags)
        pair_key_to_destinations[pair_key].add(dst)

    # Emit rows: (sources_tuple, destinations_tuple, port_tags_tuple)
    rows = []
    for (sources_frozen, port_tags), destinations in pair_key_to_destinations.items():
        dests_sorted = tuple(sorted(destinations, key=_ip_sort_key))
        rows.append((sources_frozen, dests_sorted, port_tags))

    # Sort output: by first destination, then by first source
    rows.sort(key=lambda r: (_ip_sort_key(r[1][0]), _ip_sort_key(r[0][0])))

    stats = {
        "initiator_tuples": len(initiator_tuples),
        "confirmed_tuples": len(confirmed_tuples),
        "unconfirmed_tuples": sorted(unconfirmed_tuples, key=lambda t: (
            _ip_sort_key(t[0]), _ip_sort_key(t[1]), t[2], t[3]
        )),
    }
    return rows, stats


def _ip_sort_key(ip):
    """Sort key that puts IPs in numeric order, with non-IPs at the end."""
    parts = ip.split(".")
    if len(parts) == 4:
        try:
            return tuple(int(p) for p in parts)
        except ValueError:
            pass
    return (999, 999, 999, 999, ip)


def write_xlsx(rows, stats, path, source_label=None, source_files=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Intent Table"

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="404040")
    body_font = Font(name="Calibri", size=10)
    wrap = Alignment(wrap_text=True, vertical="top")

    headers = ["Source", "Destination", "PROTOCOL [PORTS]"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    for idx, (sources, destinations, port_tags) in enumerate(rows, start=2):
        ws.cell(row=idx, column=1, value=", ".join(sources)).font = body_font
        ws.cell(row=idx, column=2, value=", ".join(destinations)).font = body_font
        ws.cell(row=idx, column=3, value=", ".join(port_tags)).font = body_font
        for col in range(1, 4):
            ws.cell(row=idx, column=col).alignment = wrap

    # Column widths tuned for readability
    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 40
    ws.freeze_panes = "A2"

    # Unconfirmed Flows sheet — initiator tuples with no matching responder.
    # These are NOT in the policy output. Engineer reviews to decide if they
    # are blocked traffic (good — leave out), aged-out responders (re-capture),
    # or one-way UDP (whitelist manually).
    unc = wb.create_sheet("Unconfirmed Flows")
    unc_headers = ["Source", "Destination", "Protocol", "Dest Port"]
    for col, h in enumerate(unc_headers, start=1):
        c = unc.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    for idx, (src, dst, proto, dport) in enumerate(stats["unconfirmed_tuples"], start=2):
        proto_name = {6: "TCP", 17: "UDP", 1: "ICMP"}.get(proto, str(proto))
        unc.cell(row=idx, column=1, value=src).font = body_font
        unc.cell(row=idx, column=2, value=dst).font = body_font
        unc.cell(row=idx, column=3, value=proto_name).font = body_font
        unc.cell(row=idx, column=4, value=dport).font = body_font
    unc.column_dimensions["A"].width = 22
    unc.column_dimensions["B"].width = 22
    unc.column_dimensions["C"].width = 10
    unc.column_dimensions["D"].width = 12
    unc.freeze_panes = "A2"

    # Provenance sheet — small audit trail
    meta = wb.create_sheet("About")
    meta_lines = [
        ("Generated by", "flow_to_hs.py"),
        ("Source input", source_label or "(stdin)"),
        ("Source file count", str(len(source_files)) if source_files else "1"),
        ("Initiator tuples (raw)", str(stats["initiator_tuples"])),
        ("  confirmed (kept)", str(stats["confirmed_tuples"])),
        ("  unconfirmed (dropped)", str(len(stats["unconfirmed_tuples"]))),
        ("Intent rows emitted", str(len(rows))),
        ("", ""),
        ("Reduction principles applied:", ""),
        ("  1. Stateful", "responder rows drive confirmation, then dropped"),
        ("  2. Global", "direction / DPU / interface context dropped"),
        ("  3. Unordered", "no rule precedence, permits only"),
        ("  4. Port-tagged", "multiple ports per (src,dst) comma-collapsed"),
        ("", ""),
        ("Bidirectional confirmation gate:", ""),
        ("  Rule", "an initiator (src,dst,proto,dport) is emitted only if"),
        ("  ", "a responder row with (src=dst, dst=src, proto=proto, sport=dport)"),
        ("  ", "exists in at least one input snapshot. Verified empirically."),
        ("", ""),
        ("Multi-snapshot union:", ""),
        ("  Behavior", "all snapshots feed one tuple stream. A tuple is confirmed"),
        ("  ", "if ANY snapshot contains a matching responder. Truly blocked"),
        ("  ", "flows stay init-only across all snapshots and are never emitted."),
        ("", ""),
        ("Grouping rule (two-pass):", ""),
        ("  Pass 1: by (dst, port-set)", "sources sharing the exact port-set against the same destination → one row"),
        ("  Pass 2: by (src-set, port-set)", "destinations sharing the same source-set and port-set → one row"),
    ]
    for i, (k, v) in enumerate(meta_lines, start=1):
        meta.cell(row=i, column=1, value=k).font = Font(name="Calibri", size=10, bold=bool(k and not v))
        meta.cell(row=i, column=2, value=v).font = Font(name="Calibri", size=10)
    meta.column_dimensions["A"].width = 35
    meta.column_dimensions["B"].width = 75

    # Source Files sheet — list every input file (only when multiple were used)
    if source_files and len(source_files) > 1:
        sf = wb.create_sheet("Source Files")
        sf_headers = ["#", "File Name", "Full Path"]
        for col, h in enumerate(sf_headers, start=1):
            c = sf.cell(row=1, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center")
        for idx, fpath in enumerate(source_files, start=2):
            sf.cell(row=idx, column=1, value=idx - 1).font = body_font
            sf.cell(row=idx, column=2, value=fpath.name).font = body_font
            sf.cell(row=idx, column=3, value=str(fpath)).font = body_font
        sf.column_dimensions["A"].width = 6
        sf.column_dimensions["B"].width = 45
        sf.column_dimensions["C"].width = 80
        sf.freeze_panes = "A2"

    wb.save(path)


def gather_input_files(directory):
    """
    Return a sorted list of capture files inside `directory`.
    Sort order = filename (ISO-8601 timestamped names sort chronologically).
    """
    p = Path(directory)
    if not p.is_dir():
        sys.exit(f"Input must be a directory: {p}")
    files = sorted(p.glob("flowcap_*.txt"))
    if not files:
        sys.exit(f"No flowcap_*.txt files found in: {p}")
    return files


# --------------------------------------------------------------------------
# HS / Hypershield CSV emitter
# --------------------------------------------------------------------------
#
# Produces the three CSVs that import_hs_policy_csv.py consumes:
#   network_objects.csv    — one row per unique source-set or destination-set
#   policies.csv           — one row per (intent_row × port_tag)
#   policy_group.csv       — single-row draft group with timestamped name
#
# Object naming scheme (deterministic, KISS — engineer renames in CSV before
# upload if they want pretty names):
#   single IP:  OBJ_<ip with dots as underscores>           e.g. OBJ_10_3_6_20
#   single CIDR: OBJ_<ip>_<mask>                            e.g. OBJ_10_3_6_0_24
#   multi-host: OBJ_<first-ip>_<count>hosts                 e.g. OBJ_10_3_6_20_2hosts
#
# An identical set always produces the same key, so dedup happens naturally.
# --------------------------------------------------------------------------

PROTO_FOR_CSV = {"TCP": "tcp", "UDP": "udp", "ICMP": "icmp"}


def _object_key_for_set(ip_set):
    """Build a deterministic, readable object key for a set of IPs/CIDRs."""
    ips = sorted(ip_set, key=_ip_sort_key)
    first = ips[0]
    # Normalize: dots and slashes both become underscores
    safe = first.replace(".", "_").replace("/", "_")
    if len(ips) == 1:
        return f"OBJ_{safe}"
    return f"OBJ_{safe}_{len(ips)}hosts"


def _parse_port_tag(tag):
    """
    Parse a port tag like 'TCP [80]' or 'UDP [53]' into (protocol, dport_str).
    The dport_str is left as-is so a range like '1024-65535' would pass through;
    flow_to_hs.py currently only emits single ports, so this is forward-safe.
    """
    m = re.match(r"^\s*(TCP|UDP|ICMP)\s*\[(.+)\]\s*$", tag)
    if not m:
        return None, None
    proto = PROTO_FOR_CSV.get(m.group(1))
    return proto, m.group(2).strip()


def _reset_output_dir(output_dir: Path):
    """
    Prepare an output directory for generated CSV files.

    The directory contents are removed, but the function refuses obviously
    dangerous targets such as filesystem roots, the user's home directory,
    and the current working directory.
    """
    import shutil

    target = output_dir.expanduser().resolve()
    protected_paths = {
        Path(target.anchor).resolve(),
        Path.home().resolve(),
        Path.cwd().resolve(),
    }
    if target in protected_paths:
        raise ValueError(f"Refusing to clear unsafe output directory: {target}")

    target.mkdir(parents=True, exist_ok=True)

    for child in target.iterdir():
        if child.is_dir():
            shutil.rmtree(child)
        else:
            child.unlink()


def write_hs_csvs(rows, output_dir: Path):
    """
    Emit network_objects.csv, policies.csv, policy_group.csv into output_dir.
    Clears output_dir contents first so each run is a fresh draft.
    """
    import csv

    _reset_output_dir(output_dir)

    # Build the unique object-set inventory across both source and destination columns
    object_specs = {}  # key -> (name, addresses_list)
    for sources, destinations, _port_tags in rows:
        for ip_set in (sources, destinations):
            key = _object_key_for_set(ip_set)
            if key not in object_specs:
                object_specs[key] = list(sorted(ip_set, key=_ip_sort_key))

    # ---- network_objects.csv ----
    netobj_path = output_dir / "network_objects.csv"
    with netobj_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow([
            "object_key", "name", "description", "object_type",
            "addresses", "vrf", "vlan", "pod_labels", "kube_namespaces",
            "existing_id", "hs_object_name",
        ])
        for key, addresses in object_specs.items():
            w.writerow([
                key, key, "",
                "NETWORK",
                ";".join(addresses),
                "", "", "", "", "", "",
            ])

    # ---- policies.csv ----
    # One policy per intent row. Multiple rules per policy when the row has
    # multiple port tags. policy_key is shared across the row's rules.
    pol_path = output_dir / "policies.csv"
    with pol_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow([
            "policy_key", "name", "description", "effect", "effect_and",
            "source_object", "destination_object", "protocol",
            "source_port", "destination_port", "policy_id",
        ])
        for idx, (sources, destinations, port_tags) in enumerate(rows, start=1):
            src_key = _object_key_for_set(sources)
            dst_key = _object_key_for_set(destinations)
            policy_key = f"POL_{idx:03d}_{src_key}_to_{dst_key}"
            policy_name = policy_key  # engineer can rename
            for tag in port_tags:
                proto, dport = _parse_port_tag(tag)
                if proto is None:
                    continue
                w.writerow([
                    policy_key, policy_name, "",
                    "permit", "nolog",
                    src_key, dst_key,
                    proto,
                    "",        # source_port: ephemeral, leave empty
                    dport,
                    "",        # policy_id: assigned by HS-MP-API
                ])

    # ---- policy_group.csv ----
    pg_path = output_dir / "policy_group.csv"
    stamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    with pg_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["name", "description"])
        w.writerow([
            f"flow_to_hs draft {stamp}",
            f"Generated by flow_to_hs.py on {stamp}",
        ])

    return netobj_path, pol_path, pg_path, len(object_specs)


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument(
        "input_dir",
        help="Directory containing flowcap_*.txt files written by flow_collector.py",
    )
    ap.add_argument("-o", "--output", required=True, help="Output .xlsx path")
    ap.add_argument(
        "--hs-csv-dir", default=None,
        help="If set, also emit Hypershield import CSVs into this directory "
             "(network_objects.csv, policies.csv, policy_group.csv). Directory "
             "contents are cleared on each run.",
    )
    args = ap.parse_args()

    files = gather_input_files(args.input_dir)
    text = "\n".join(f.read_text(errors="replace") for f in files)
    label = f"{len(files)} files from {args.input_dir}"

    tuples = list(parse_dump(text))
    if not tuples:
        sys.exit("No flow rows parsed. Check the capture files contain dpctl show flow output.")

    rows, stats = reduce_flows(tuples)
    write_xlsx(rows, stats, args.output, source_label=label, source_files=files)

    initiators = sum(1 for *_, role in tuples if role.lower() == "initiator")
    responders = len(tuples) - initiators
    print(f"Inputs:   {len(files)} file(s) from {args.input_dir}")
    print(f"Parsed:   {len(tuples)} flow rows ({initiators} init / {responders} resp)")
    print(f"Tuples:   {stats['initiator_tuples']} seen, "
          f"{stats['confirmed_tuples']} confirmed, "
          f"{len(stats['unconfirmed_tuples'])} dropped")
    print(f"Output:   {len(rows)} intent rows  →  {args.output}")

    if args.hs_csv_dir:
        hs_dir = Path(args.hs_csv_dir)
        netobj_path, pol_path, pg_path, obj_count = write_hs_csvs(rows, hs_dir)
        print(f"HS CSVs: {obj_count} network objects  →  {hs_dir}/")


if __name__ == "__main__":
    main()
